"""
Générateur de templates Excel
"""

from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from loguru import logger

from backend.models.template_config import TemplateConfig


class ExcelTemplateGenerator:
    """Génère des templates Excel pré-configurés"""
    
    def __init__(self, config: TemplateConfig):
        self.config = config
    
    def generate(
        self,
        source_file: Optional[Path] = None,
        output_dir: Path = Path("."),
        create_new: bool = False
    ) -> Path:
        """
        Génère un template Excel.
        
        Args:
            source_file: Fichier Excel source (ou None)
            output_dir: Dossier de sortie
            create_new: Créer un nouveau fichier même si source fournie
            
        Returns:
            Chemin du fichier généré
        """
        output_path = output_dir / "master.xlsx"
        
        if not create_new and source_file and source_file.exists():
            logger.info(f"Copie du fichier Excel source : {source_file}")
            import shutil
            shutil.copy2(source_file, output_path)
            return output_path
        
        logger.info("Création d'un nouveau template Excel")
        
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Créer les feuilles
        self._create_balises_sheet(wb)
        self._create_Boucles_sheet(wb)
        self._create_table_sheet(wb)
        
        # Sauvegarder
        wb.save(output_path)
        logger.success(f"Template Excel créé : {output_path}")
        
        return output_path
    
    def _create_balises_sheet(self, wb: openpyxl.Workbook) -> None:
        """Crée la feuille Balises avec les paramètres"""
        ws = wb.create_sheet("Balises", 0)
        
        # En-têtes
        headers = ["Balise", "Description", "Valeur", "Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Ajouter les paramètres
        for row, param in enumerate(self.config.parameters, 2):
            ws.cell(row, 1, param.balise_ppt)
            ws.cell(row, 2, param.description or "")
            ws.cell(row, 3, str(param.default) if param.default else "")
            ws.cell(row, 4, param.type)
        
        # Créer un tableau structuré
        tab = Table(
            displayName="Balises",
            ref=f"A1:D{len(self.config.parameters) + 1}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def _create_Boucles_sheet(self, wb: openpyxl.Workbook) -> None:
        """Crée la feuille Boucles"""
        ws = wb.create_sheet("Boucles")
        
        # Paramètres de filtrage
        ws['A1'] = "Paramètres de filtrage"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['B3'] = "Distributeur"
        ws['B4'] = "Sous-marque"
        
        ws['C3'] = "Leclerc"
        ws['C4'] = "BOMBAY"
        
        # Tableau Loop si nécessaire
        if self.config.loops:
            ws['A6'] = "Tableau Loop"
            ws['A6'].font = Font(bold=True, size=14)
            
            ws['A8'] = "ID"
            ws['B8'] = "Itération"
            ws['C8'] = "Nombre de tests"
            
            for col in ['A8', 'B8', 'C8']:
                ws[col].font = Font(bold=True, color="FFFFFF")
                ws[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for row, loop in enumerate(self.config.loops, 9):
                ws.cell(row, 1, loop.loop_id)
                ws.cell(row, 2, 1)
                ws.cell(row, 3, 0)
            
            # Créer tableau structuré Loop
            tab = Table(
                displayName="Loop",
                ref=f"A8:C{8 + len(self.config.loops)}"
            )
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
    
    def _create_table_sheet(self, wb: openpyxl.Workbook) -> None:
        """Crée la feuille Table pour les données"""
        ws = wb.create_sheet("Table")
        
        ws['A1'] = "Feuille de données"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Les tableaux de données seront injectés ici"